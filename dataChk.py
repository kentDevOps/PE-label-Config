from sqliteProcess import *
import os,sys
import pandas as pd
from dataProcess import *
import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
def normalize_text(x):
    if pd.isna(x):
        return ''              # chuyển NaN => '' để NaN không bị tính là khác
    s = str(x)
    # Normalization Unicode (NFKC giúp gộp ký tự dạng kết hợp)
    s = unicodedata.normalize('NFKC', s)
    # Thay non-breaking space bằng space thường
    s = s.replace('\u00A0', ' ')
    # Remove zero-width spaces and BOM
    s = re.sub(r'[\u200B-\u200D\uFEFF]', '', s)
    # Remove CR/LF, chuyển nhiều dòng thành 1 dòng
    s = s.replace('\r', ' ').replace('\n', ' ')
    # Trim và thay nhiều space bằng 1 space
    s = s.strip()
    s = re.sub(r'\s+', ' ', s)
    return s
def to_mau(excel_path):
    wb = load_workbook(excel_path)
    ws = wb['result']

    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        diff = ws[f"E{row}"].value  # cột Khác nhau?
        if diff:
            ws[f"C{row}"].fill = fill_red  # Value_std
            ws[f"D{row}"].fill = fill_red  # Value_if
def export_compare_result(rs_df,project):
    strAbsPath = os.path.abspath(sys.argv[0])
    strDirPath = os.path.dirname(strAbsPath)
    output_file = f'Result_Compare_Ifuse_Config_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    if not os.path.exists(os.path.join(strDirPath,'report')):
        os.makedirs(os.path.join(strDirPath,'report'))
    if not os.path.exists(os.path.join(strDirPath,'report',project,'Ifuse Check')):
        os.makedirs(os.path.join(strDirPath,'report',project,'Ifuse Check'))
    output_file = os.path.join(os.path.join(strDirPath,'report',project,'Ifuse Check'),output_file)
    rs_df.to_excel(output_file, index=False, sheet_name='result')
    return output_file
def compare_ifuse_std(std_df,ifuse_df):
    # Gộp hai bảng theo SKU và Type
    compare_df = pd.merge(std_df, ifuse_df, on=['SKU', 'Type'], how='outer', suffixes=('_std', '_ifuse'))

    # Chuẩn hóa text trước khi so sánh
    compare_df['Value_std_clean'] = compare_df['Value_std'].apply(normalize_text)
    compare_df['Value_if_clean']  = compare_df['Value_ifuse'].apply(normalize_text)

    # So sánh cột Value đã clean
    compare_df['Result'] = compare_df['Value_std_clean'] != compare_df['Value_if_clean']
    compare_df['Result'] = compare_df['Result'].map({True: 'Different', False: 'Same'})
    # (Tuỳ chọn) giữ lại các cột chính cho dễ đọc
    compare_df = compare_df[['SKU', 'Type', 'Value_std', 'Value_ifuse', 'Result']]

    print(compare_df)
    return compare_df
def sort_dict(dict_data):
    # Danh sách thể hiện thứ tự mong muốn
    thu_tu = ['Des', 'Brand', 'UPC', 'Rap', 'Bob','sap','color','price','mrp','qr-code']

    # Tạo “bản đồ thứ tự” (mapping) từ tên → vị trí
    # enumerate() gán chỉ số cho từng phần tử trong 'thu_tu'
    # kết quả: {'xoài': 0, 'ổi': 1, 'chuối': 2, 'cam': 3, 'bưởi': 4}
    order_map = {key: i for i, key in enumerate(thu_tu)}

    # Sắp xếp các key trong 'data' theo vị trí trong 'order_map'
    # - sorted(data) duyệt qua danh sách key trong dict
    # - key=lambda x: order_map.get(x, len(thu_tu)):
    #     -> lấy chỉ số thứ tự của key x trong order_map
    #     -> nếu không tìm thấy thì trả về len(thu_tu) (đưa ra cuối)
    sorted_keys = sorted(dict_data, key=lambda x: order_map.get(x, len(thu_tu)))

    # Tạo dict mới dựa trên thứ tự đã sắp xếp ở trên
    # Duyệt từng key trong sorted_keys, lấy value tương ứng từ 'data'
    sorted_dict = {k: dict_data[k] for k in sorted_keys}

    # In kết quả cuối cùng
    print(sorted_dict)
    return sorted_dict
def gen_std(project,product,list_sku,suffix):
    strAbsPath = os.path.abspath(sys.argv[0])
    strDirPath = os.path.dirname(strAbsPath)
    chk_mandatory_folder(strAbsPath,strDirPath)
    # return list address of file
    folder_path = os.path.join(strDirPath,"database")
    list_data_file = [os.path.join(folder_path,item) for item in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path,item))]
    print(list_data_file) 
    # chk whether data file exists or not
    config_path = ''
    mapping_path = ''
    keyFile_path = ''
    for item in list_data_file:
        if project=='MT5':
            if 'VN MP Delivery Tracker' in os.path.basename(item):
                mapping_path = item
        elif project=='BZ5':
            if 'INT-VN-BZ5 MP POR' in os.path.basename(item):
                mapping_path = item  
        elif project=='FL5':
            if 'FL5 POR Packout' in os.path.basename(item):
                mapping_path = item 
        if 'config' in os.path.basename(item):
            config_path = item
        elif 'keyFile' in os.path.basename(item):
            keyFile_path = item    
    if project == "MT5":
        pd_mapping = pd.read_excel(mapping_path,sheet_name='VN MP SKU Mapping Table',dtype={'EAN': str,'UPC碼' : str})
    elif project == "FL5":
        pd_mapping = pd.read_excel(mapping_path,sheet_name='FL5_POR_包裝對照表 (Mapping Table)',header=10,dtype={'UPC Code-12': str,'EAN Code-13' : str})
    elif project == "BZ5":
        pd_mapping = pd.read_excel(mapping_path,sheet_name='BZ5 MP Control table_ VN',header=6,dtype={'UPC Code-12\n(except EMEA)': str,'EAN Code-13\n(EMEA only)' : str,'SKU ID (C-SKU for ATT only)': str})
        pd_des = pd.read_excel(mapping_path,sheet_name='POR Config List')
    new_rows_df = []
    fixed_headers = ['SKU', 'Type', 'Value']
    main_df = pd.DataFrame(columns=fixed_headers)
    dict_region = {}
    for item in list_sku:
        print(item)
        if suffix != '':
            suff_item = item + '-' + suffix
        else:
            suff_item = item
        arr_mapping = list()
        #Loc trong mapping de compare
        mapp_des = mapp_upc = map_bob = mapp_rap = mapp_brand = ''
        mapp_mrp=mapp_price=mapp_qr=''
        mapp_sap=mapp_carton_color = ''
        mapp_region = ''
        if project == "MT5":
            df_loc_sku = pd_mapping[pd_mapping.iloc[:,1] == item]
            df_loc_sku.columns = df_loc_sku.columns.str.replace('\xa0', '', regex=False)
            mapp_region = df_loc_sku.iloc[0,20].strip()
            mapp_des = df_loc_sku.iloc[0:5]
            arr_mapping = filter_mapping_config(mapp_region,df_loc_sku,item,config_path,project,product)
        elif project == 'FL5':
            df_loc_mapping = pd_mapping[pd_mapping.iloc[:,4] == item]
            print(df_loc_mapping)
            df_loc_mapping.columns = df_loc_mapping.columns.str.replace('\xa0', '', regex=False)
            mapp_region = df_loc_mapping.iloc[0,5].strip()
            mapp_des = df_loc_mapping.iloc[0,28]
            arr_mapping = filter_mapping_config_FL(mapp_region,df_loc_mapping,item,config_path,project,product)
        elif project == 'BZ5':
            df_loc_mapping = pd_mapping[pd_mapping.iloc[:,6] == item]
            df_loc_mapping.columns = df_loc_mapping.columns.str.replace('\xa0', '', regex=False)
            mapp_region = df_loc_mapping.iloc[0,7].strip()
            mapp_des = df_loc_mapping.iloc[0,26]
            arr_mapping = filter_mapping_config_BZ(mapp_region,df_loc_mapping,item,config_path,mapp_des,project,product)   
        print(arr_mapping)
        print(f'------------------Query all Information  of sku : {item} from Keyconfig file-------------------')    
        mapp_region = region_convert(mapp_region)
        list_data = list()
        Des_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'Des')
        brand_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'Brand')
        upc_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'UPC')
        rap_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'Rap')
        bob_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'Bob')
        list_data = [Des_list,brand_list,upc_list,rap_list,bob_list]
        if mapp_region in ['ATT','TMO']:
            sap_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'sap')
            color_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'color')
            list_data.append(sap_list)
            list_data.append(color_list)
        elif mapp_region in ['IN']: 
            price_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'price')
            mrp_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'mrp')
            qr_list = gen_dict_standard(mapp_region,arr_mapping,suff_item,'qr-code')
            list_data.append(price_list)
            list_data.append(mrp_list)
            list_data.append(qr_list)
        print(list_data)
        new_rows_df = pd.DataFrame(list_data)
        main_df = pd.concat([main_df , new_rows_df], ignore_index=True)
        print(main_df)
        dict_region[item] = mapp_region
    return main_df,dict_region
def main_chk(project,product,keyconfig_path):
    strAbsPath = os.path.abspath(sys.argv[0])
    strDirPath = os.path.dirname(strAbsPath)
    chk_mandatory_folder(strAbsPath,strDirPath)
    # return list address of file
    folder_path = os.path.join(strDirPath,"database")
    list_data_file = [os.path.join(folder_path,item) for item in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path,item))]
    print(list_data_file) 
    # chk whether data file exists or not
    config_path = ''
    mapping_path = ''
    keyFile_path = ''
    for item in list_data_file:
        if project=='MT5':
            if 'VN MP Delivery Tracker' in os.path.basename(item):
                mapping_path = item
        elif project=='BZ5':
            if 'INT-VN-BZ5 MP POR' in os.path.basename(item):
                mapping_path = item  
        elif project=='FL5':
            if 'FL5 POR Packout' in os.path.basename(item):
                mapping_path = item 
        if 'config' in os.path.basename(item):
            config_path = item
        elif 'keyFile' in os.path.basename(item):
            keyFile_path = item
    chk_mandatary_file(config_path,mapping_path,keyFile_path)
    df_keyConfig = pd.read_excel(keyconfig_path,dtype={'key描述': str,'key值' : str})
    if project == "MT5":
        pd_mapping = pd.read_excel(mapping_path,sheet_name='VN MP SKU Mapping Table',dtype={'EAN': str,'UPC碼' : str})
    elif project == "FL5":
        pd_mapping = pd.read_excel(mapping_path,sheet_name='FL5_POR_包裝對照表 (Mapping Table)',header=10,dtype={'UPC Code-12': str,'EAN Code-13' : str})
    elif project == "BZ5":
        pd_mapping = pd.read_excel(mapping_path,sheet_name='BZ5 MP Control table_ VN',header=6,dtype={'UPC Code-12\n(except EMEA)': str,'EAN Code-13\n(EMEA only)' : str,'SKU ID (C-SKU for ATT only)': str})
        pd_des = pd.read_excel(mapping_path,sheet_name='POR Config List')
    unique_Sku = df_keyConfig['料號'].unique()
    new_rows_df = []
    fixed_headers = ['SKU', 'Type', 'Value', 'Confirm', 'Key']
    main_df = pd.DataFrame(columns=fixed_headers)
    suff = ''
    for item in unique_Sku:
        print(f'------------------Start checking sku : {item}-------------------')
        if len(item)>10 and 'CLNR' not in item and 'IQ' not in item:
            mapp = item[:10]
            suff = item.split('-')[-1]
        else:
            mapp = item
        arr_mapping = list()
        #Loc trong mapping de compare
        mapp_des = mapp_upc = map_bob = mapp_rap = mapp_brand = ''
        mapp_mrp=mapp_price=mapp_qr=''
        mapp_sap=mapp_carton_color = ''
        mapp_region = ''
        if project == "MT5":
            df_loc_sku = pd_mapping[pd_mapping.iloc[:,1] == mapp]
            df_loc_sku.columns = df_loc_sku.columns.str.replace('\xa0', '', regex=False)
            mapp_region = df_loc_sku.iloc[0,20].strip()
            mapp_des = df_loc_sku.iloc[0:5]
            arr_mapping = filter_mapping_config(mapp_region,df_loc_sku,mapp,config_path,project,product)
        elif project == 'FL5':
            df_loc_mapping = pd_mapping[pd_mapping.iloc[:,4] == mapp]
            print(df_loc_mapping)
            df_loc_mapping.columns = df_loc_mapping.columns.str.replace('\xa0', '', regex=False)
            mapp_region = df_loc_mapping.iloc[0,5].strip()
            mapp_des = df_loc_mapping.iloc[0,28]
            arr_mapping = filter_mapping_config_FL(mapp_region,df_loc_mapping,mapp,config_path,project,product)
        elif project == 'BZ5':
            df_loc_mapping = pd_mapping[pd_mapping.iloc[:,6] == mapp]
            df_loc_mapping.columns = df_loc_mapping.columns.str.replace('\xa0', '', regex=False)
            mapp_region = df_loc_mapping.iloc[0,7].strip()
            mapp_des = df_loc_mapping.iloc[0,26]
            arr_mapping = filter_mapping_config_BZ(mapp_region,df_loc_mapping,mapp,config_path,mapp_des,project,product)
        print(arr_mapping)
        print(f'------------------Query all Information  of sku : {item} from Keyconfig file-------------------')
        print(df_keyConfig)
        list_std = list()
        des_value,des_confirm,des_keyname = filter_standard_config(df_keyConfig,item,'Description')
        brand_value,brand_confirm,brand_keyname = filter_standard_config(df_keyConfig,item,'Brand')
        upc_value,upc_confirm,upc_keyname = filter_standard_config(df_keyConfig,item,'UPC')
        rap_value,rap_confirm,rap_keyname = filter_standard_config(df_keyConfig,item,'R-Label')
        bob_value,bob_confirm,bob_keyname = filter_standard_config(df_keyConfig,item,'B-Label')
        #chk cac gia tri trong file keyConfig vowi Mapping
        list_compare_des = compare_keyConfig_mapping(item,des_value,des_confirm,des_keyname,arr_mapping,'Des')
        list_compare_brand = compare_keyConfig_mapping(item,brand_value,brand_confirm,brand_keyname,arr_mapping,'Brand')
        list_compare_upc = compare_keyConfig_mapping(item,upc_value,upc_confirm,upc_keyname,arr_mapping,'UPC')
        list_compare_rap = compare_keyConfig_mapping(item,rap_value,rap_confirm,rap_keyname,arr_mapping,'Rap')
        list_compare_bob = compare_keyConfig_mapping(item,bob_value,bob_confirm,bob_keyname,arr_mapping,'Bob')
        #gop cac dict vao lam thanh 1 list 
        list_compare = [list_compare_des,list_compare_brand,list_compare_upc,list_compare_rap,list_compare_bob]
        new_rows_df = pd.DataFrame(list_compare)
        print(new_rows_df)
        region = region_convert(mapp_region)
        if region == 'ATT' or region == 'TMO':
            sap_value,sap_confirm,sap_keyname = filter_standard_config(df_keyConfig,item,'SAP')
            color_value,color_confirm,color_keyname = filter_standard_config(df_keyConfig,item,'Carton_Color')
            list_compare_sap = compare_keyConfig_mapping(item,sap_value,sap_confirm,sap_keyname,arr_mapping,'sap')
            list_compare_color = compare_keyConfig_mapping(item,color_value,color_confirm,color_keyname,arr_mapping,'color')
            list_att = [list_compare_sap,list_compare_color]
            new_rows_df = pd.concat([new_rows_df, pd.DataFrame(list_att)], ignore_index=True)
        elif region == 'IN':
            mrp_value,mrp_confirm,mrp_keyname = filter_standard_config(df_keyConfig,item,'M-Label')
            qr_value,qr_confirm,qr_keyname = filter_standard_config(df_keyConfig,item,'QR_CODE')
            price_value,price_confirm,price_keyname = filter_standard_config(df_keyConfig,item,'Price')

            list_compare_mrp = compare_keyConfig_mapping(item,mrp_value,mrp_confirm,mrp_keyname,arr_mapping,'mrp')
            list_compare_qr = compare_keyConfig_mapping(item,qr_value,qr_confirm,qr_keyname,arr_mapping,'gpn')
            list_compare_price = compare_keyConfig_mapping(item,price_value,price_confirm,price_keyname,arr_mapping,'price')
            list_att = [list_compare_mrp,list_compare_qr,list_compare_price]
            new_rows_df = pd.concat([new_rows_df, pd.DataFrame(list_att)], ignore_index=True)
        print(new_rows_df)
        main_df = pd.concat([main_df , new_rows_df], ignore_index=True)
    print(main_df)
    return main_df